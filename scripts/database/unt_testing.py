import os
import sys
from pathlib import Path
from dotenv import load_dotenv
import openpyxl

load_dotenv()
BASE_PATH = Path(os.getenv("GS_BASE_PATH"))
RUTA_MACRO = BASE_PATH / "04-AuditoriaCalidadVisualAtributiva.xlsm"
RUTA_PORTS = BASE_PATH / "01-Bulk export of ports.xlsx"

def normalizar_id(valor):
    if valor is None or valor == "":
        return ""
    if isinstance(valor, (int, float)):
        return str(int(valor))
    return str(valor).strip()

# Cargar IDs OM
ids_om = set()
try:
    wb = openpyxl.load_workbook(RUTA_MACRO, read_only=True, keep_vba=True)
    if "OM" in wb.sheetnames:
        ws_om = wb["OM"]
        for row in ws_om.iter_rows(min_row=2, values_only=True):
            if row[0]:
                ids_om.add(normalizar_id(row[0]))
    wb.close()
    print(f"IDs OM cargados: {len(ids_om)}")
except Exception as e:
    print(f"Error cargando OM: {e}")

# Leer puertos
try:
    wb = openpyxl.load_workbook(RUTA_PORTS, read_only=True)
    ws = wb["Port"]
    # Buscar columna ID
    id_col = None
    for idx, cell in enumerate(ws[1], start=0):
        if cell.value and str(cell.value).strip().lower() == "id":
            id_col = idx
            break
    if id_col is None:
        print("No se encontró columna 'id'")
        sys.exit(1)

    total = 0
    en_om = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if id_col < len(row):
            idv = normalizar_id(row[id_col])
            if idv:
                total += 1
                if idv in ids_om:
                    en_om += 1
    print(f"Total puertos: {total}")
    print(f"Puertos en OM: {en_om}")
    print(f"Puertos NO en OM: {total - en_om}")
    wb.close()
except Exception as e:
    print(f"Error: {e}")