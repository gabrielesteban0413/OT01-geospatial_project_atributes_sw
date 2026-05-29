import openpyxl
from pathlib import Path
import pandas as pd
import os
from dotenv import load_dotenv

load_dotenv()

BASE_PATH_STR = os.getenv("GS_BASE_PATH")
if not BASE_PATH_STR:
    raise ValueError("Variable GS_BASE_PATH no encontrada en .env")

BASE_PATH = Path(BASE_PATH_STR)

ARCHIVO_ORIGINAL = BASE_PATH / "01-Bulk export of ports.xlsx"
ARCHIVO_NUEVO = BASE_PATH / "01-Bulk export of ports1.xlsx"
OUTPUT_EXCEL = BASE_PATH / "comparacion_ports.xlsx"

COLUMNAS_INTERES = ['Id', 'Cambio Service Manager', 'Id Servicio']

def extraer_datos(ruta):
    wb = openpyxl.load_workbook(ruta, read_only=True)
    ws = wb.active
    
    headers = {}
    for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), 1):
        if cell.value:
            headers[cell.value] = idx
    
    col_indices = []
    for col in COLUMNAS_INTERES:
        if col in headers:
            col_indices.append((col, headers[col]))
    
    datos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        fila = {}
        for col_name, idx in col_indices:
            val = row[idx-1] if idx-1 < len(row) else None
            fila[col_name] = str(val).strip() if val is not None else ''
        if fila.get('Id'):
            datos.append(fila)
    
    wb.close()
    return datos

def comparar_ports():
    print("Leyendo archivo original...")
    original = extraer_datos(ARCHIVO_ORIGINAL)
    
    print("Leyendo archivo nuevo...")
    nuevo = extraer_datos(ARCHIVO_NUEVO)
    
    print(f"Original: {len(original)} filas")
    print(f"Nuevo: {len(nuevo)} filas")
    
    ids_original = {row['Id'] for row in original}
    ids_nuevo = {row['Id'] for row in nuevo}
    
    ids_perdidos = ids_original - ids_nuevo
    ids_nuevos = ids_nuevo - ids_original
    
    print(f"\nIDs perdidos: {len(ids_perdidos)}")
    print(f"IDs nuevos: {len(ids_nuevos)}")
    
    datos_perdidos = [row for row in original if row['Id'] in ids_perdidos]
    datos_nuevos = [row for row in nuevo if row['Id'] in ids_nuevos]
    
    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        pd.DataFrame({
            'Concepto': ['Original', 'Nuevo', 'IDs perdidos', 'IDs nuevos'],
            'Valor': [len(original), len(nuevo), len(ids_perdidos), len(ids_nuevos)]
        }).to_excel(writer, sheet_name='Resumen', index=False)
        
        if datos_perdidos:
            pd.DataFrame(datos_perdidos).to_excel(writer, sheet_name='IDs_perdidos', index=False)
        
        if datos_nuevos:
            pd.DataFrame(datos_nuevos).to_excel(writer, sheet_name='IDs_nuevos', index=False)
    
    print(f"\nReporte guardado en: {OUTPUT_EXCEL}")
    
    if datos_perdidos:
        print(f"\nIDs ELIMINADOS (primeros 10):")
        for row in datos_perdidos[:10]:
            print(f"  ID: {row['Id']} | Cambio Service Manager: {row['Cambio Service Manager']} | Id Servicio: {row['Id Servicio']}")
    
    if datos_nuevos:
        print(f"\nIDs NUEVOS AGREGADOS (primeros 10):")
        for row in datos_nuevos[:10]:
            print(f"  ID: {row['Id']} | Cambio Service Manager: {row['Cambio Service Manager']} | Id Servicio: {row['Id Servicio']}")

if __name__ == "__main__":
    comparar_ports()