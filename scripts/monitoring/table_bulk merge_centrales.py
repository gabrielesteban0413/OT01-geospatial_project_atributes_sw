import pandas as pd
import re
from pathlib import Path
from collections import defaultdict
import warnings
import openpyxl
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
import traceback
import os
from dotenv import load_dotenv

load_dotenv()

warnings.filterwarnings('ignore')

BASE_PATH_STR = os.getenv("GS_BASE_PATH")
if not BASE_PATH_STR:
    raise ValueError("Variable GS_BASE_PATH no encontrada en .env")

BASE_PATH = Path(BASE_PATH_STR)

RUTA_CENTRALES = BASE_PATH / "01-Bulk Merge_Centrales.xlsx"
RUTA_PORTS = BASE_PATH / "01-Bulk export of ports.xlsx"
RUTA_ANILLOS_CAMBIO = BASE_PATH / "03-anillos-cambio.xlsx"
RUTA_BASE_EC = Path(os.getenv("GS_BASE_EQUIPO-CENTRAL", ""))
RUTA_BASE_IS = Path(os.getenv("GS_BASE_CARPETA12", ""))


HOJA_ENTRADA_EC = "EC"
HOJA_SALIDA_OP = "OP"
HOJA_SALIDA_OT = "OT"
HOJA_SALIDA_IS = "IS"
HOJA_ANILLOS = "AnilloC"
HOJA_OM = "OM"

EXTENSIONES_VALIDAS = {'.xlsx', '.xlsm', '.xls'}
FILA_INICIO = 7
TITULOS_PROHIBIDOS = ['OPTICOS', 'ODF', 'CASSETERA', 'POS', 'EQUIPO', 'INTERFAZ', 'PUERTO']

EC_COL_ARCHIVO = "ARCHIVO"
EC_COL_EQUIPO = "EQUIPO CENTRAL"
EC_COL_PUERTO = "PUERTO CENTRAL"
EC_COL_ODF = "ODF"
EC_COL_CASSETERA = "CASSETERA"
EC_COL_POS = "POS"
EC_COL_POS2 = "POS2"
EC_COL_ID_PORT = "ID_PORT_CABECERA"

OP_COL_ARCHIVO = "ARCHIVO"
OP_COL_EQUIPO = "EQUIPO CENTRAL"
OP_COL_PUERTO = "PUERTO CENTRAL"
OP_COL_ODF = "ODF"
OP_COL_CASSETERA = "CASSETERA"
OP_COL_POS_ORIG = "POS"
OP_COL_POS_TOKEN = "POS DIV"
OP_COL_ID_CABECERA = "ID_PORT_CABECERA"
OP_COL_ID_PORT_PACHEO = "ID_PORT_PACHEO"

OT_COL_ID = "ID"
OT_COL_FIBRA = "FIBRA"
OT_COL_NOMBRE_ODF = "NOMBRE ODF"
OT_COL_DESC_PORT = "DESCRIPTION PORT"
OT_COL_DESC_SHELF = "DESCRIPCION SHELF"

IS_REAL_EC = "EC"
IS_REAL_PORT_EC = "PORT_EC"
IS_REAL_PATCHEO = "PATCHEO"
IS_REAL_FIBRA = "FIBRA"
IS_REAL_PORT_OT = "PORT_OT"
IS_REAL_HILO = "HILO"
IS_REAL_ANILLO_ORIGEN = "ANILLO ORIGEN"
IS_REAL_BUL = "BUL"
IS_REAL_COMENTARIOS = "COMENTARIOS"

IS_COL_FIBRA = "FIBRA"
IS_COL_EQUIPO = "EC"
IS_COL_INTERFAZ = "PORT_EC"
IS_COL_PATCHEO = "PATCHEO"
IS_COL_ODF_CALLE = "PORT_OT"
IS_COL_HILO = "HILO"
IS_COL_ANILLO_ORIGEN = "ANILLO ORIGEN"
IS_COL_CASSETERA = "CASSETERA"
IS_COL_OP = "ODF P"
IS_COL_BANDEJA = "BANDEJA"
IS_COL_ID_CABECERA = "ID_PORT_CABECERA"
IS_COL_ID_PORT_PACHEO = "ID_PORT_PACHEO"

COL_ID_PORT_ODF_TRONCAL = "ID_PORT_ODF-TRONCAL"
COL_DESC_PORT_OT = "DESCRIPTION PORT OT"
COL_DESC_SHELF_OT = "DESCRIPCION SHELF OT"

PORTS_COL_ID = "ID"
PORTS_COL_DESC = "DESCRIPTION"
PORTS_COL_DESC2 = "DESCRIPTION (2)"
PORTS_COL_DESC3 = "DESCRIPTION (3)"
PORTS_COL_ID_BAY = "ID BAY"

POSIBLES_ID = ["Id", "ID", "id"]
POSIBLES_DESC = ["Description", "descripcion"]
POSIBLES_DESC2 = ["Description (2)", "descripcion (2)"]
POSIBLES_DESC3 = ["Description (3)", "descripcion (3)"]
POSIBLES_ID_BAY = ["Id Bay", "ID Bay", "id bay", "Bay Id", "BAY ID","Bay.id"]
POSIBLES_DESC_SHELF = ["Description Shelf", "description shelf", "DESCRIPTION SHELF"]

def cargar_cables_omitidos(ruta_centrales):
    cables_omitidos = set()
    try:
        if not ruta_centrales.exists():
            return cables_omitidos
        try:
            df_om = pd.read_excel(ruta_centrales, sheet_name=HOJA_OM, dtype=str)
            col_fibra = None
            for col in df_om.columns:
                if col.upper() == "FIBRA":
                    col_fibra = col
                    break
            if col_fibra is not None:
                for valor in df_om[col_fibra].dropna():
                    fibra = str(valor).strip()
                    if fibra:
                        cables_omitidos.add(fibra)
        except ValueError:
            pass
    except Exception as e:
        print(f"Error al cargar cables omitidos: {e}", file=sys.stderr)
    return cables_omitidos

def es_fila_valida(valores):
    for val in valores:
        if pd.notna(val):
            val_str = str(val).strip().upper()
            if any(titulo in val_str for titulo in TITULOS_PROHIBIDOS):
                return False
    return True

def extraer_partes_puerto(puerto_str):
    if pd.isna(puerto_str):
        return None, None, None
    limpio = re.sub(r'^[A-Z]+[\s-]?', '', str(puerto_str))
    partes = limpio.strip().split('/')
    if len(partes) == 3:
        try:
            return int(partes[0]), int(partes[1]), int(partes[2])
        except ValueError:
            return None, None, None
    return None, None, None

def normalizar_columna(df, posibles_nombres):
    columnas_lower = {col.strip().lower(): col for col in df.columns}
    for nombre in posibles_nombres:
        if nombre.lower() in columnas_lower:
            return columnas_lower[nombre.lower()]
    return None

def get_engine(archivo):
    ext = archivo.suffix.lower()
    if ext == '.xls':
        return 'calamine'
    elif ext in ('.xlsx', '.xlsm'):
        return 'openpyxl'
    return None

def extraer_equipo_central(archivo):
    try:
        engine = get_engine(archivo)
        df = pd.read_excel(archivo, header=None, usecols="B", nrows=2, engine=engine, dtype=str)
        if not df.empty and len(df) >= 2:
            valor = df.iat[1, 0]
            if pd.notna(valor):
                valor = str(valor).strip()
                if valor.startswith("EQUIPO "):
                    valor = valor[7:]
                if ' ' in valor:
                    valor = valor.split(' ', 1)[0]
                return valor
    except Exception:
        pass
    return ""

def extraer_ultimo_numero_odf(valor):
    if pd.isna(valor):
        return valor
    valor_str = str(valor).strip()
    if valor_str.isdigit():
        return valor_str
    match = re.search(r'ODF\s*(\d+)', valor_str, re.IGNORECASE)
    if match:
        return match.group(1)
    numeros = re.findall(r'\d+', valor_str)
    if numeros:
        return numeros[-1]
    return valor_str

def limpiar_puerto_completo(valor):
    if pd.isna(valor):
        return valor
    cadena = re.sub(r'[^0-9/-]', '', str(valor))
    if cadena.endswith('/'):
        cadena = cadena[:-1]
    partes = cadena.split('/')
    partes_limpias = [p.lstrip('0') or '0' for p in partes]
    return '/'.join(partes_limpias)

def limpiar_puerto_is(valor):
    if pd.isna(valor):
        return valor
    cadena = str(valor).strip()
    if cadena.endswith('/'):
        cadena = cadena[:-1]
    partes = cadena.split('/')
    partes_limpias = [p.lstrip('0') or '0' for p in partes]
    return '/'.join(partes_limpias)

def extraer_anillo_desde_origen(valor):
    if pd.isna(valor):
        return None
    cadena = str(valor).strip()
    if cadena == '#N/A' or cadena == 'N/A' or cadena == '':
        return None
    patrones = [
        r'(ME[A-Z0-9]+)',
        r'([A-Z]{2,}[A-Z0-9]{3,})',
    ]
    for patron in patrones:
        matches = re.findall(patron, cadena, re.IGNORECASE)
        if matches:
            for match in matches:
                if len(match) >= 6:
                    if not re.match(r'^\d+\.\d+\.\d+\.\d+$', match):
                        return match.upper()
    partes = re.split(r'[_/]', cadena)
    for parte in reversed(partes):
        parte = re.sub(r'\(.*\)', '', parte)
        if parte and len(parte) >= 6 and not parte.isdigit():
            if not re.match(r'^\d+\.\d+\.\d+\.\d+$', parte):
                return parte.upper()
    return None

def cargar_datos_anillos(ruta_anillos):
    try:
        if not ruta_anillos.exists():
            return {}
        df_anillos = pd.read_excel(ruta_anillos, sheet_name=HOJA_ANILLOS)
        df_anillos.columns = [c.upper().strip() for c in df_anillos.columns]
        columnas_requeridas = ['CABLE', 'ANILLO', 'ANILLO ACTUAL', 'ILUMINACION DE ANILLO', 'FECHA CAMBIO', 'IGUALAR']
        for col in columnas_requeridas:
            if col not in df_anillos.columns:
                return {}
        df_anillos['FECHA CAMBIO'] = pd.to_datetime(df_anillos['FECHA CAMBIO'], errors='coerce')
        def formatear_fecha(fecha):
            if pd.notna(fecha):
                return fecha.strftime('%Y-%m-%d')
            return ''
        df_anillos['FECHA CAMBIO_STR'] = df_anillos['FECHA CAMBIO'].apply(formatear_fecha)
        anillos_dict = {}
        for _, row in df_anillos.iterrows():
            cable = str(row['CABLE']).strip() if pd.notna(row['CABLE']) else ''
            anillo = str(row['ANILLO']).strip() if pd.notna(row['ANILLO']) else ''
            anillo_actual = str(row['ANILLO ACTUAL']).strip() if pd.notna(row['ANILLO ACTUAL']) else ''
            iluminacion = str(row['ILUMINACION DE ANILLO']).strip() if pd.notna(row['ILUMINACION DE ANILLO']) else ''
            fecha_cambio = row['FECHA CAMBIO_STR']
            igualar = str(row['IGUALAR']).strip().upper() if pd.notna(row['IGUALAR']) else ''
            registro = {
                'cable': cable,
                'anillo': anillo,
                'anillo_actual': anillo_actual,
                'iluminacion': iluminacion,
                'fecha_cambio': fecha_cambio,
                'igualar': igualar
            }
            if anillo:
                anillos_dict[anillo] = registro
            if anillo_actual and anillo_actual != anillo:
                anillos_dict[anillo_actual] = registro
        return anillos_dict
    except Exception:
        return {}

def procesar_cambio_anillo(fibra, anillo, anillos_dict):
    if pd.isna(anillo) or str(anillo).strip() == '':
        return "SIN ANILLO"
    anillo_str = str(anillo).strip()
    fibra_str = str(fibra).strip() if pd.notna(fibra) else ''
    if anillo_str in anillos_dict:
        registro = anillos_dict[anillo_str]
        if registro['cable'] and fibra_str and registro['cable'] != fibra_str:
            if registro['igualar'] == 'SI':
                pass
            else:
                return "SIN COINCIDENCIA"
        if registro['anillo_actual'] == anillo_str:
            return f"V|{registro['anillo_actual']}|{registro['iluminacion']}|{registro['fecha_cambio']}"
        if registro['anillo'] == anillo_str:
            anillo_a_mostrar = registro['anillo_actual'] if registro['anillo_actual'] else registro['anillo']
            return f"F|{anillo_a_mostrar}|{registro['iluminacion']}|{registro['fecha_cambio']}"
    return "SIN COINCIDENCIA"

def validar_anillos_por_fibra_y_hilo(df_is):
    if df_is.empty or 'ANILLO' not in df_is.columns or IS_COL_FIBRA not in df_is.columns or IS_COL_HILO not in df_is.columns:
        return df_is
    df_validado = df_is.copy()
    for (fibra, hilo), grupo in df_validado.groupby([IS_COL_FIBRA, IS_COL_HILO]):
        anillos_unicos = grupo['ANILLO'].dropna().unique()
        if len(anillos_unicos) > 1:
            mascara_error = grupo.index
            df_validado.loc[mascara_error, 'SINCRONIZA'] = df_validado.loc[mascara_error, 'SINCRONIZA'].apply(
                lambda x: f"{x} - ERROR: Hilos con diferentes anillos" if pd.notna(x) and str(x).strip() != '' 
                else "ERROR: Hilos con diferentes anillos"
            )
    return df_validado

def leer_un_archivo_ec(archivo):
    try:
        equipo = extraer_equipo_central(archivo)
        engine = get_engine(archivo)
        df_header = pd.read_excel(
            archivo,
            header=None,
            nrows=10,
            usecols="A:F",
            engine=engine,
            dtype=str
        )
        col_map = {'puerto': None, 'odf': None, 'cassetera': None, 'pos': None}
        palabras_clave = {
            'puerto': ['puerto', 'interfaz', 'port'],
            'odf': ['odf'],
            'cassetera': ['cassetera', 'cassette'],
            'pos': ['pos', 'posición']
        }
        for row_idx in range(min(10, len(df_header))):
            row = df_header.iloc[row_idx].astype(str).str.lower().str.strip()
            for col_idx, val in enumerate(row):
                if pd.isna(val) or val == 'nan':
                    continue
                for key, keywords in palabras_clave.items():
                    if col_map[key] is None and any(kw in val for kw in keywords):
                        col_map[key] = col_idx
                        break
        if None in col_map.values():
            df_raw = pd.read_excel(
                archivo,
                skiprows=FILA_INICIO - 1,
                header=None,
                usecols="A:F",
                engine=engine,
                dtype=str
            )
            if df_raw.empty:
                return pd.DataFrame()
            num_cols = df_raw.shape[1]
            if num_cols < 6:
                for i in range(num_cols, 6):
                    df_raw[i] = pd.NA
            primera_fila_con_datos = None
            for i in range(len(df_raw)):
                if pd.notna(df_raw.iloc[i, 1]):
                    primera_fila_con_datos = df_raw.iloc[i]
                    break
            if primera_fila_con_datos is None:
                return pd.DataFrame()
            col_a_tiene_dato = pd.notna(primera_fila_con_datos.iloc[0]) and str(primera_fila_con_datos.iloc[0]).strip() != ''
            if col_a_tiene_dato:
                valor_col_d = primera_fila_con_datos.iloc[3] if len(primera_fila_con_datos) > 3 else None
                if pd.notna(valor_col_d) and str(valor_col_d).strip() != '' and str(valor_col_d).strip().replace('.','',1).isdigit():
                    indices = [2, 3, 4, 5]
                else:
                    indices = [2, 4, 5, 6] if 6 < num_cols else [2, 3, 4, 5]
            else:
                valor_col_c = primera_fila_con_datos.iloc[2] if len(primera_fila_con_datos) > 2 else None
                if pd.notna(valor_col_c) and str(valor_col_c).strip() != '' and str(valor_col_c).strip().replace('.','',1).isdigit():
                    indices = [1, 2, 3, 4]
                else:
                    indices = [1, 3, 4, 5]
            indices_validos = [i for i in indices if i < num_cols]
            if len(indices_validos) < 4:
                df = pd.DataFrame()
                for idx, col_name in zip(indices, [EC_COL_PUERTO, EC_COL_ODF, EC_COL_CASSETERA, EC_COL_POS]):
                    if idx < num_cols:
                        df[col_name] = df_raw[idx]
                    else:
                        df[col_name] = pd.NA
            else:
                df = df_raw[indices_validos].copy()
                df.columns = [EC_COL_PUERTO, EC_COL_ODF, EC_COL_CASSETERA, EC_COL_POS]
            df[EC_COL_PUERTO] = df[EC_COL_PUERTO].apply(limpiar_puerto_completo)
            df['_ODF_original'] = df[EC_COL_ODF]
            df[EC_COL_ODF] = df[EC_COL_ODF].apply(extraer_ultimo_numero_odf)
            df[EC_COL_POS2] = pd.NA
            df.insert(0, EC_COL_EQUIPO, equipo)
            df.insert(0, EC_COL_ARCHIVO, archivo.name)
            df.columns = [c.upper() for c in df.columns]
            return df
        df_raw = pd.read_excel(
            archivo,
            skiprows=FILA_INICIO - 1,
            header=None,
            usecols="A:F",
            engine=engine,
            dtype=str
        )
        if df_raw.empty:
            return pd.DataFrame()
        columnas = {}
        for key, col_idx in col_map.items():
            if col_idx is not None and col_idx < df_raw.shape[1]:
                columnas[key] = df_raw[col_idx]
            else:
                columnas[key] = pd.Series(pd.NA, index=df_raw.index)
        df = pd.DataFrame({
            EC_COL_PUERTO: columnas['puerto'],
            EC_COL_ODF: columnas['odf'],
            EC_COL_CASSETERA: columnas['cassetera'],
            EC_COL_POS: columnas['pos']
        })
        df[EC_COL_PUERTO] = df[EC_COL_PUERTO].apply(limpiar_puerto_completo)
        df['_ODF_original'] = df[EC_COL_ODF]
        df[EC_COL_ODF] = df[EC_COL_ODF].apply(extraer_ultimo_numero_odf)
        df[EC_COL_POS2] = pd.NA
        df.insert(0, EC_COL_EQUIPO, equipo)
        df.insert(0, EC_COL_ARCHIVO, archivo.name)
        df.columns = [c.upper() for c in df.columns]
        return df
    except Exception:
        return pd.DataFrame()

def generar_hoja_ec(ruta_carpeta):
    if not ruta_carpeta.exists():
        return pd.DataFrame()
    archivos = [a for a in ruta_carpeta.glob("*") if a.suffix.lower() in EXTENSIONES_VALIDAS]
    if not archivos:
        return pd.DataFrame()
    dataframes = []
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(leer_un_archivo_ec, archivo): archivo for archivo in archivos}
        for idx, future in enumerate(as_completed(futures), 1):
            try:
                df = future.result()
                if not df.empty:
                    dataframes.append(df)
            except Exception:
                pass
    if dataframes:
        df_ec = pd.concat(dataframes, ignore_index=True)
        df_ec[EC_COL_EQUIPO] = df_ec.apply(
            lambda row: Path(row[EC_COL_ARCHIVO]).stem 
            if pd.isna(row[EC_COL_EQUIPO]) or str(row[EC_COL_EQUIPO]).strip() == '' 
            else row[EC_COL_EQUIPO],
            axis=1
        )
        return df_ec
    return pd.DataFrame()

def cargar_y_preprocesar_ports(ruta):
    try:
        df = pd.read_excel(ruta, sheet_name="PortCard", dtype=str)
        col_id = normalizar_columna(df, POSIBLES_ID)
        col_desc = normalizar_columna(df, POSIBLES_DESC)
        col_desc2 = normalizar_columna(df, POSIBLES_DESC2)
        col_desc3 = normalizar_columna(df, POSIBLES_DESC3)
        col_id_bay = normalizar_columna(df, POSIBLES_ID_BAY)
        if not all([col_id, col_desc, col_desc2, col_desc3, col_id_bay]):
            raise ValueError("No se encontraron las columnas necesarias en el archivo de ports")
        df = df.rename(columns={
            col_id: PORTS_COL_ID,
            col_desc: PORTS_COL_DESC,
            col_desc2: PORTS_COL_DESC2,
            col_desc3: PORTS_COL_DESC3,
            col_id_bay: PORTS_COL_ID_BAY
        })
        id_a_id_bay = {}
        for _, row in df.iterrows():
            if pd.notna(row[PORTS_COL_ID]):
                id_a_id_bay[str(row[PORTS_COL_ID]).strip()] = str(row[PORTS_COL_ID_BAY]).strip() if pd.notna(row[PORTS_COL_ID_BAY]) else None
        df['slot_extraido'] = df[PORTS_COL_DESC2].str.extract(r'Slot\s*(\d+)', expand=False).astype(float).fillna(-1).astype(int)
        df.loc[df['slot_extraido'] == -1, 'slot_extraido'] = None
        tipo_num = df[PORTS_COL_DESC].str.extract(r'^(XG|G)-(\d+)$', expand=True)
        df['tipo_puerto'] = tipo_num[0].str.upper()
        df['numero_puerto'] = pd.to_numeric(tipo_num[1], errors='coerce').astype('Int64')
        mask_validas = df['slot_extraido'].notna() & df['tipo_puerto'].notna() & df['numero_puerto'].notna()
        df_validos = df[mask_validas].copy()
        df_validos['es_xg'] = df_validos['tipo_puerto'] == 'XG'
        tiene_xg_por_slot = df_validos.groupby([PORTS_COL_DESC3, 'slot_extraido'])['es_xg'].any().to_dict()
        tiene_xg_por_slot = {(k[0], k[1]): v for k, v in tiene_xg_por_slot.items()}
        dict_busqueda = defaultdict(list)
        for _, row in df_validos.iterrows():
            clave = (row[PORTS_COL_DESC3], row['slot_extraido'], row['numero_puerto'])
            dict_busqueda[clave].append((row['tipo_puerto'], row[PORTS_COL_ID]))
        df['bandeja'] = df[PORTS_COL_DESC2].str.extract(r'BANDEJA[^A-Za-z]*([A-Za-z])', expand=False).str.upper()
        tiene_op = df[PORTS_COL_DESC3].str.contains('OP', na=False, case=False)
        df_pacheo_potencial = df[df['bandeja'].notna() & tiene_op].copy()
        pacheo_dict = {}
        for _, row in df_pacheo_potencial.iterrows():
            op_key = str(row[PORTS_COL_DESC3]).strip() if pd.notna(row[PORTS_COL_DESC3]) else ''
            token = str(row[PORTS_COL_DESC]).strip() if pd.notna(row[PORTS_COL_DESC]) else ''
            if op_key and token:
                key = (row['bandeja'], op_key, token)
                if key not in pacheo_dict:
                    pacheo_dict[key] = row[PORTS_COL_ID]
        return dict_busqueda, tiene_xg_por_slot, pacheo_dict, id_a_id_bay
    except Exception as e:
        print(f"Error en cargar_y_preprocesar_ports({ruta}): {e}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        raise

def buscar_id_port_backbone(fila, dict_busqueda, tiene_xg_por_slot):
    equipo = str(fila[EC_COL_EQUIPO]).strip() if pd.notna(fila[EC_COL_EQUIPO]) else ''
    puerto = fila[EC_COL_PUERTO]
    if not equipo or pd.isna(puerto):
        return None
    slot, medio, numero = extraer_partes_puerto(puerto)
    if slot is None or medio is None or numero is None:
        return None
    clave = (equipo, slot, numero)
    candidatos = dict_busqueda.get(clave, [])
    if not candidatos:
        return None
    tiene_xg = tiene_xg_por_slot.get((equipo, slot), False)
    for tipo, id_val in candidatos:
        if tipo == 'XG' and medio == 0:
            return id_val
        if tipo == 'G':
            if tiene_xg and medio == 1:
                return id_val
            if not tiene_xg and medio == 0:
                return id_val
    return None

def buscar_id_backbone_directo(equipo, puerto, dict_busqueda, tiene_xg_por_slot):
    if pd.isna(equipo) or pd.isna(puerto):
        return None
    equipo = str(equipo).strip()
    puerto = str(puerto).strip()
    slot, medio, numero = extraer_partes_puerto(puerto)
    if slot is None or medio is None or numero is None:
        return None
    clave = (equipo, slot, numero)
    candidatos = dict_busqueda.get(clave, [])
    if not candidatos:
        return None
    tiene_xg = tiene_xg_por_slot.get((equipo, slot), False)
    for tipo, id_val in candidatos:
        if tipo == 'XG' and medio == 0:
            return id_val
        if tipo == 'G':
            if tiene_xg and medio == 1:
                return id_val
            if not tiene_xg and medio == 0:
                return id_val
    return None

def generar_hojas_op_y_ot(df_cabecera):
    try:
        dict_busqueda, tiene_xg_por_slot, pacheo_dict, id_a_id_bay = cargar_y_preprocesar_ports(RUTA_PORTS)
    except Exception as e:
        print(f"Error al cargar ports en generar_hojas_op_y_ot: {e}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        return df_cabecera, pd.DataFrame(), pd.DataFrame()
    if EC_COL_ID_PORT not in df_cabecera.columns:
        df_cabecera[EC_COL_ID_PORT] = df_cabecera.apply(
            lambda fila: buscar_id_port_backbone(fila, dict_busqueda, tiene_xg_por_slot), axis=1
        )
    df_cabecera['_archivo'] = df_cabecera[EC_COL_ARCHIVO]
    df_cabecera['_equipo'] = df_cabecera[EC_COL_EQUIPO].astype(str).str.strip()
    df_cabecera['_odf'] = df_cabecera[EC_COL_ODF].astype(str).str.strip()
    df_cabecera['_cassetera'] = df_cabecera[EC_COL_CASSETERA].astype(str).str.strip().str.upper()
    df_cabecera['_puerto_central'] = df_cabecera[EC_COL_PUERTO].apply(limpiar_puerto_completo)
    df_cabecera['_id_backbone'] = df_cabecera[EC_COL_ID_PORT]
    df_cabecera['_op_key'] = df_cabecera.apply(
        lambda row: f"OP{row['_equipo']}_{row['_odf']}" if row['_equipo'] and row['_odf'] else None, axis=1
    )
    def procesar_valor_fila(row, col_name):
        valor = row[col_name]
        if pd.isna(valor) or str(valor).strip() == '':
            return []
        valor_normalizado = str(valor).replace('.', ',')
        tokens = [t.strip() for t in valor_normalizado.split(',') if t.strip()]
        registros = []
        for token in tokens:
            id_port_pacheo = None
            if row['_cassetera'] and row['_op_key'] and token:
                clave_buscada = (row['_cassetera'], row['_op_key'], token)
                id_port_pacheo = pacheo_dict.get(clave_buscada)
            id_cabecera_asignado = None
            if row['_id_backbone'] and id_port_pacheo:
                id_bay_backbone = id_a_id_bay.get(str(row['_id_backbone']))
                id_bay_pacheo = id_a_id_bay.get(str(id_port_pacheo))
                if id_bay_backbone is not None and id_bay_pacheo is not None and str(id_bay_backbone) == str(id_bay_pacheo):
                    id_cabecera_asignado = row['_id_backbone']
            registros.append({
                OP_COL_ARCHIVO: row['_archivo'],
                OP_COL_EQUIPO: row['_equipo'],
                OP_COL_PUERTO: row['_puerto_central'],
                OP_COL_ODF: row['_odf'],
                OP_COL_CASSETERA: row['_cassetera'],
                OP_COL_POS_ORIG: valor,
                OP_COL_POS_TOKEN: token,
                OP_COL_ID_CABECERA: id_cabecera_asignado,
                OP_COL_ID_PORT_PACHEO: id_port_pacheo
            })
        return registros
    df_cabecera['_registros_pos'] = df_cabecera.apply(lambda r: procesar_valor_fila(r, EC_COL_POS), axis=1)
    df_cabecera['_registros_pos2'] = df_cabecera.apply(lambda r: procesar_valor_fila(r, EC_COL_POS2), axis=1)
    all_registros = []
    for _, row in df_cabecera.iterrows():
        registros = row['_registros_pos'] + row['_registros_pos2']
        if not registros:
            all_registros.append({
                OP_COL_ARCHIVO: row['_archivo'],
                OP_COL_EQUIPO: row['_equipo'],
                OP_COL_PUERTO: row['_puerto_central'],
                OP_COL_ODF: row['_odf'],
                OP_COL_CASSETERA: row['_cassetera'],
                OP_COL_POS_ORIG: None,
                OP_COL_POS_TOKEN: 'SIN_POS',
                OP_COL_ID_CABECERA: None,
                OP_COL_ID_PORT_PACHEO: None
            })
        else:
            all_registros.extend(registros)
    df_op = pd.DataFrame(all_registros)
    df_op.columns = [c.upper() for c in df_op.columns]
    try:
        df_port_sheet = pd.read_excel(RUTA_PORTS, sheet_name="Port", dtype=str)
        df_ot = procesar_odf_troncal(df_port_sheet)
        if not df_ot.empty:
            df_ot.columns = [c.upper() for c in df_ot.columns]
    except Exception as e:
        print(f"Error al procesar OT desde Port: {e}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        df_ot = pd.DataFrame()
    return df_cabecera, df_op, df_ot

def procesar_odf_troncal(df_port):
    try:
        col_id = normalizar_columna(df_port, POSIBLES_ID)
        col_desc = normalizar_columna(df_port, POSIBLES_DESC)
        col_desc_shelf = normalizar_columna(df_port, POSIBLES_DESC_SHELF)
        if not all([col_id, col_desc, col_desc_shelf]):
            return pd.DataFrame(columns=[OT_COL_ID, OT_COL_FIBRA, OT_COL_NOMBRE_ODF, OT_COL_DESC_PORT, OT_COL_DESC_SHELF])
        mask = df_port[col_desc_shelf].astype(str).str.startswith('OT', na=False)
        df_filtrado = df_port[mask].copy()
        records = []
        for _, row in df_filtrado.iterrows():
            desc_shelf = str(row[col_desc_shelf]).strip()
            fibra = ''
            nombre_odf = ''
            desc_shelf_rest = ''
            if desc_shelf.startswith('OT'):
                resto = desc_shelf[2:]
                if '_' in resto:
                    partes_underscore = resto.split('_', 1)
                    fibra = partes_underscore[0]
                    resto2 = partes_underscore[1]
                    if ' ' in resto2:
                        partes_espacio = resto2.split(' ', 1)
                        nombre_odf = partes_espacio[0]
                        desc_shelf_rest = partes_espacio[1]
                    else:
                        nombre_odf = resto2
                else:
                    if ' ' in resto:
                        partes_espacio = resto.split(' ', 1)
                        fibra = partes_espacio[0]
                        desc_shelf_rest = partes_espacio[1]
                    else:
                        fibra = resto
            records.append({
                OT_COL_ID: row[col_id],
                OT_COL_FIBRA: fibra,
                OT_COL_NOMBRE_ODF: nombre_odf,
                OT_COL_DESC_PORT: row[col_desc],
                OT_COL_DESC_SHELF: desc_shelf_rest
            })
        return pd.DataFrame(records)
    except Exception as e:
        print(f"Error en procesar_odf_troncal: {e}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        return pd.DataFrame(columns=[OT_COL_ID, OT_COL_FIBRA, OT_COL_NOMBRE_ODF, OT_COL_DESC_PORT, OT_COL_DESC_SHELF])

def leer_un_archivo_is(archivo):
    """
    Lee un archivo de cable (IS) y retorna solo las filas donde la columna BUL sea igual a 1.
    """
    fibra_archivo = archivo.stem
    # Omitir archivos temporales
    if archivo.name.startswith('~$'):
        return []
    fibra = archivo.stem
    filas_archivo = []
    try:
        ext = archivo.suffix.lower()
        if ext in ('.xlsx', '.xlsm'):
            wb = openpyxl.load_workbook(archivo, data_only=True)
            sheet_names = wb.sheetnames
            target_sheet = None
            for sn in sheet_names:
                if '-' in sn:
                    target_sheet = sn
                    break
            if target_sheet is not None:
                ws = wb[target_sheet]
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                if header_row[0] is None:
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                        if any(cell is not None for cell in row):
                            header_row = row
                            break
                col_indices = {}
                for col_idx, val in enumerate(header_row):
                    if val is None:
                        continue
                    val_clean = str(val).strip()
                    if val_clean == "EC":
                        col_indices['equipo'] = col_idx
                    elif val_clean == "PORT_EC":
                        col_indices['interfaz'] = col_idx
                    elif val_clean == "PATCHEO":
                        col_indices['patcheo'] = col_idx
                    elif val_clean == "FIBRA":
                        col_indices['fibra'] = col_idx
                    elif val_clean == "PORT_OT":
                        col_indices['odf_calle'] = col_idx
                    elif val_clean == "HILO":
                        col_indices['hilo'] = col_idx
                    elif val_clean == "ANILLO ORIGEN":
                        col_indices['anillo_origen'] = col_idx
                    elif val_clean == "BUL":
                        col_indices['bul'] = col_idx
                if all(k in col_indices for k in ['equipo', 'interfaz', 'patcheo']):
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
                        if ws.row_dimensions[row_idx].hidden:
                            continue
                        if any(cell is not None for cell in row):
                            # Obtener el valor de BUL
                            bul_val = row[col_indices['bul']] if 'bul' in col_indices and col_indices['bul'] < len(row) else None
                            # Verificar si BUL es 1 (como cadena o número)
                            bul_ok = False
                            if bul_val is not None:
                                bul_str = str(bul_val).strip()
                                if bul_str == '1' or (isinstance(bul_val, (int, float)) and bul_val == 1):
                                    bul_ok = True
                            if not bul_ok:
                                continue  # omitir filas donde BUL no sea 1
                            equipo_val = row[col_indices['equipo']] if col_indices['equipo'] < len(row) else None
                            interfaz_val = row[col_indices['interfaz']] if col_indices['interfaz'] < len(row) else None
                            patcheo_val = row[col_indices['patcheo']] if col_indices['patcheo'] < len(row) else None
                            odf_calle_val = row[col_indices.get('odf_calle')] if col_indices.get('odf_calle') is not None and col_indices['odf_calle'] < len(row) else None
                            hilo_val = row[col_indices.get('hilo')] if col_indices.get('hilo') is not None and col_indices['hilo'] < len(row) else None
                            anillo_origen_val = row[col_indices.get('anillo_origen')] if col_indices.get('anillo_origen') is not None and col_indices['anillo_origen'] < len(row) else None
                            if interfaz_val is not None:
                                interfaz_val = limpiar_puerto_completo(interfaz_val)
                            anillo_val = extraer_anillo_desde_origen(anillo_origen_val)
                            if equipo_val is not None or interfaz_val is not None or patcheo_val is not None:
                                filas_archivo.append({
                                    IS_COL_FIBRA: fibra,
                                    IS_COL_EQUIPO: equipo_val,
                                    IS_COL_INTERFAZ: interfaz_val,
                                    IS_COL_PATCHEO: patcheo_val,
                                    IS_COL_ODF_CALLE: odf_calle_val,
                                    IS_COL_HILO: hilo_val,
                                    IS_COL_ANILLO_ORIGEN: anillo_origen_val,
                                    'ANILLO': anillo_val
                                })
                else:
                    # Si faltan columnas esenciales, no se incluye ninguna fila (o se podría incluir una vacía)
                    pass
        elif ext == '.xls':
            engine = 'calamine'
            xl = pd.ExcelFile(archivo, engine=engine)
            sheet_names = xl.sheet_names
            target_sheet = None
            for sn in sheet_names:
                if '-' in sn:
                    target_sheet = sn
                    break
            if target_sheet is not None:
                df = pd.read_excel(archivo, sheet_name=target_sheet, engine=engine, dtype=str)
                df.columns = [str(col).strip() for col in df.columns]
                col_equipo = None
                col_interfaz = None
                col_patcheo = None
                col_odf_calle = None
                col_hilo = None
                col_anillo_origen = None
                col_bul = None
                for col in df.columns:
                    col_clean = col.strip()
                    if col_clean == "EC":
                        col_equipo = col
                    elif col_clean == "PORT_EC":
                        col_interfaz = col
                    elif col_clean == "PATCHEO":
                        col_patcheo = col
                    elif col_clean == "PORT_OT":
                        col_odf_calle = col
                    elif col_clean == "HILO":
                        col_hilo = col
                    elif col_clean == "ANILLO ORIGEN":
                        col_anillo_origen = col
                    elif col_clean == "BUL":
                        col_bul = col
                if col_equipo and col_interfaz and col_patcheo:
                    # Filtrar por BUL == 1
                    if col_bul is not None:
                        df = df[df[col_bul].astype(str).str.strip() == '1']
                    else:
                        # Si no existe columna BUL, no se incluye ninguna fila (o se puede cambiar a df vacío)
                        df = pd.DataFrame()
                    if not df.empty:
                        subdf = df[[col_equipo, col_interfaz, col_patcheo]].copy()
                        subdf.columns = [IS_COL_EQUIPO, IS_COL_INTERFAZ, IS_COL_PATCHEO]
                        if col_odf_calle:
                            subdf[IS_COL_ODF_CALLE] = df[col_odf_calle]
                        else:
                            subdf[IS_COL_ODF_CALLE] = None
                        if col_hilo:
                            subdf[IS_COL_HILO] = df[col_hilo]
                        else:
                            subdf[IS_COL_HILO] = None
                        if col_anillo_origen:
                            subdf[IS_COL_ANILLO_ORIGEN] = df[col_anillo_origen]
                            subdf['ANILLO'] = subdf[IS_COL_ANILLO_ORIGEN].apply(extraer_anillo_desde_origen)
                        else:
                            subdf[IS_COL_ANILLO_ORIGEN] = None
                            subdf['ANILLO'] = None
                        subdf[IS_COL_FIBRA] = fibra
                        for _, row in subdf.iterrows():
                            interfaz_limpia = limpiar_puerto_completo(row[IS_COL_INTERFAZ])
                            filas_archivo.append({
                                IS_COL_FIBRA: row[IS_COL_FIBRA],
                                IS_COL_EQUIPO: row[IS_COL_EQUIPO],
                                IS_COL_INTERFAZ: interfaz_limpia,
                                IS_COL_PATCHEO: row[IS_COL_PATCHEO],
                                IS_COL_ODF_CALLE: row[IS_COL_ODF_CALLE],
                                IS_COL_HILO: row[IS_COL_HILO],
                                IS_COL_ANILLO_ORIGEN: row[IS_COL_ANILLO_ORIGEN] if col_anillo_origen else None,
                                'ANILLO': row['ANILLO'] if col_anillo_origen else None
                            })
    except Exception as e:
        # Si ocurre cualquier error, se omite el archivo (no se agregan filas)
        pass
    return filas_archivo

def actualizar_ids_desde_ec(df_is, df_ec):
    ec_lookup = {}
    col_equipo_ec = None
    col_puerto_ec = None
    col_id_ec = None
    for col in df_ec.columns:
        col_upper = col.upper()
        if 'EQUIPO' in col_upper and 'CENTRAL' in col_upper:
            col_equipo_ec = col
        elif 'PUERTO' in col_upper and 'CENTRAL' in col_upper:
            col_puerto_ec = col
        elif 'ID_PORT_CABECERA' in col_upper:
            col_id_ec = col
    if not all([col_equipo_ec, col_puerto_ec, col_id_ec]):
        return df_is, 0
    for _, row in df_ec.iterrows():
        equipo = str(row[col_equipo_ec]).strip() if pd.notna(row[col_equipo_ec]) else ''
        puerto = str(row[col_puerto_ec]).strip() if pd.notna(row[col_puerto_ec]) else ''
        puerto_normalizado = limpiar_puerto_completo(puerto)
        id_val = row[col_id_ec] if pd.notna(row[col_id_ec]) else ''
        if equipo and puerto_normalizado and id_val:
            id_str = str(id_val).strip()
            if 'e' in id_str.lower():
                try:
                    id_str = str(int(float(id_str)))
                except:
                    pass
            ec_lookup[(equipo, puerto_normalizado)] = id_str
    col_equipo_is = None
    col_puerto_is = None
    col_id_is = None
    col_sincroniza_is = None
    for col in df_is.columns:
        col_upper = col.upper()
        if col_upper == 'EC':
            col_equipo_is = col
        elif col_upper == 'PORT_EC':
            col_puerto_is = col
        elif 'ID_PORT_CABECERA' in col_upper:
            col_id_is = col
        elif 'SINCRONIZA' in col_upper:
            col_sincroniza_is = col
    if not all([col_equipo_is, col_puerto_is, col_id_is]):
        return df_is, 0
    actualizados = 0
    for idx, row in df_is.iterrows():
        equipo = str(row[col_equipo_is]).strip() if pd.notna(row[col_equipo_is]) else ''
        puerto = str(row[col_puerto_is]).strip() if pd.notna(row[col_puerto_is]) else ''
        puerto_normalizado = limpiar_puerto_completo(puerto)
        id_actual = str(row[col_id_is]).strip() if pd.notna(row[col_id_is]) else ''
        id_vacio = (id_actual == '' or id_actual == 'nan' or id_actual == 'None')
        if id_vacio:
            clave = (equipo, puerto_normalizado)
            if clave in ec_lookup:
                nuevo_id = ec_lookup[clave]
                df_is.at[idx, col_id_is] = nuevo_id
                actualizados += 1
                if col_sincroniza_is:
                    sincroniza_actual = str(row[col_sincroniza_is]) if pd.notna(row[col_sincroniza_is]) else ''
                    if 'F' in sincroniza_actual or 'V' not in sincroniza_actual:
                        df_is.at[idx, col_sincroniza_is] = "V - usando EC (ID respaldo)"
    return df_is, actualizados

def generar_hoja_is(ruta_cables, df_op, df_ot=None):
    """
    Genera la hoja IS a partir de los archivos de cable.
    Solo se incluyen filas donde la columna BUL sea igual a 1.
    """
    anillos_dict = cargar_datos_anillos(RUTA_ANILLOS_CAMBIO)
    try:
        dict_busqueda, tiene_xg_por_slot, pacheo_dict, id_a_id_bay = cargar_y_preprocesar_ports(RUTA_PORTS)
    except Exception:
        return pd.DataFrame(columns=[IS_COL_FIBRA, IS_COL_EQUIPO, IS_COL_INTERFAZ, IS_COL_PATCHEO, 
                                      IS_COL_ODF_CALLE, IS_COL_HILO, IS_COL_OP, IS_COL_BANDEJA, 
                                      IS_COL_CASSETERA, IS_COL_ID_CABECERA, IS_COL_ID_PORT_PACHEO, 
                                      COL_ID_PORT_ODF_TRONCAL, COL_DESC_PORT_OT, COL_DESC_SHELF_OT, 
                                      'SINCRONIZA', IS_COL_ANILLO_ORIGEN, 'ANILLO', 'CAMBIO'])
    op_ids = defaultdict(set)
    op_existe = set()
    if df_op is not None and not df_op.empty:
        df_op.columns = [c.upper() for c in df_op.columns]
        for _, row in df_op.iterrows():
            ec = str(row.get(OP_COL_EQUIPO, '')).strip()
            puerto = str(row.get(OP_COL_PUERTO, '')).strip()
            id_pacheo = str(row.get(OP_COL_ID_PORT_PACHEO, '')).strip()
            if ec and puerto:
                clave = (ec, puerto)
                op_existe.add(clave)
                if id_pacheo and id_pacheo != 'None':
                    op_ids[clave].add(id_pacheo)
    ot_mapping = None
    ot_base = None
    ot_desc_port = None
    ot_desc_shelf = None
    if df_ot is not None and not df_ot.empty:
        df_ot.columns = [c.upper() for c in df_ot.columns]
        ot_mapping = {}
        ot_base = {}
        ot_desc_port = {}
        ot_desc_shelf = {}
        for _, row in df_ot.iterrows():
            fibra = str(row[OT_COL_FIBRA]).strip()
            nombre_odf = str(row[OT_COL_NOMBRE_ODF]).strip()
            port = str(row[OT_COL_DESC_PORT]).strip()
            id_val = str(row[OT_COL_ID]).strip()
            desc_shelf = str(row[OT_COL_DESC_SHELF]).strip()
            if fibra and nombre_odf and port and id_val:
                ot_mapping[(fibra, nombre_odf, port)] = id_val
                ot_desc_port[(fibra, nombre_odf, port)] = port
                ot_desc_shelf[(fibra, nombre_odf, port)] = desc_shelf
            if fibra and nombre_odf and desc_shelf:
                match = re.search(r'(\d+)/', desc_shelf)
                if match:
                    base = int(match.group(1))
                    key = (fibra, nombre_odf)
                    if key not in ot_base:
                        ot_base[key] = base
    if not ruta_cables.exists():
        return pd.DataFrame(columns=[IS_COL_FIBRA, IS_COL_EQUIPO, IS_COL_INTERFAZ, IS_COL_PATCHEO, 
                                      IS_COL_ODF_CALLE, IS_COL_HILO, IS_COL_OP, IS_COL_BANDEJA, 
                                      IS_COL_CASSETERA, IS_COL_ID_CABECERA, IS_COL_ID_PORT_PACHEO, 
                                      COL_ID_PORT_ODF_TRONCAL, COL_DESC_PORT_OT, COL_DESC_SHELF_OT, 
                                      'SINCRONIZA', IS_COL_ANILLO_ORIGEN, 'ANILLO', 'CAMBIO'])
    archivos = [f for f in ruta_cables.glob("*") if f.is_file() and f.suffix.lower() in EXTENSIONES_VALIDAS and not f.name.startswith('~$')]
    if not archivos:
        return pd.DataFrame(columns=[IS_COL_FIBRA, IS_COL_EQUIPO, IS_COL_INTERFAZ, IS_COL_PATCHEO, 
                                      IS_COL_ODF_CALLE, IS_COL_HILO, IS_COL_OP, IS_COL_BANDEJA, 
                                      IS_COL_CASSETERA, IS_COL_ID_CABECERA, IS_COL_ID_PORT_PACHEO, 
                                      COL_ID_PORT_ODF_TRONCAL, COL_DESC_PORT_OT, COL_DESC_SHELF_OT, 
                                      'SINCRONIZA', IS_COL_ANILLO_ORIGEN, 'ANILLO', 'CAMBIO'])
    filas_totales = []
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(leer_un_archivo_is, archivo): archivo for archivo in archivos}
        for idx, future in enumerate(as_completed(futures), 1):
            try:
                filas = future.result()
                filas_totales.extend(filas)
            except Exception:
                pass
    if not filas_totales:
        return pd.DataFrame(columns=[IS_COL_FIBRA, IS_COL_EQUIPO, IS_COL_INTERFAZ, IS_COL_PATCHEO, 
                                      IS_COL_ODF_CALLE, IS_COL_HILO, IS_COL_OP, IS_COL_BANDEJA, 
                                      IS_COL_CASSETERA, IS_COL_ID_CABECERA, IS_COL_ID_PORT_PACHEO, 
                                      COL_ID_PORT_ODF_TRONCAL, COL_DESC_PORT_OT, COL_DESC_SHELF_OT, 
                                      'SINCRONIZA', IS_COL_ANILLO_ORIGEN, 'ANILLO', 'CAMBIO'])
    df_final = pd.DataFrame(filas_totales)
    
    nuevas_filas = []
    for idx, (_, fila) in enumerate(df_final.iterrows(), start=1):
        fibra = fila[IS_COL_FIBRA]
        equipo = str(fila[IS_COL_EQUIPO]).strip() if pd.notna(fila[IS_COL_EQUIPO]) else ''
        interfaz = str(fila[IS_COL_INTERFAZ]).strip() if pd.notna(fila[IS_COL_INTERFAZ]) else ''
        patcheo = fila[IS_COL_PATCHEO]
        odf_calle = fila[IS_COL_ODF_CALLE]
        hilo = fila[IS_COL_HILO]
        anillo_origen = fila.get(IS_COL_ANILLO_ORIGEN, None)
        anillo = fila.get('ANILLO', None)
        cambio = procesar_cambio_anillo(fibra, anillo, anillos_dict)
        id_backbone = buscar_id_backbone_directo(equipo, interfaz, dict_busqueda, tiene_xg_por_slot)
        
        # Valores comunes
        id_odf_troncal = None
        desc_port_ot = None
        desc_shelf_ot = None
        if ot_mapping and ot_base and pd.notna(fibra) and pd.notna(odf_calle) and pd.notna(hilo):
            fibra_str = str(fibra).strip()
            odf_calle_str = str(odf_calle).strip()
            hilo_str = str(hilo).strip()
            if hilo_str.isdigit():
                hilo_int = int(hilo_str)
                base_key = (fibra_str, odf_calle_str)
                if base_key in ot_base:
                    base_val = ot_base[base_key]
                    port = (hilo_int - base_val) + 1
                    port_str = str(port)
                    id_odf_troncal = ot_mapping.get((fibra_str, odf_calle_str, port_str))
                    desc_port_ot = ot_desc_port.get((fibra_str, odf_calle_str, port_str))
                    desc_shelf_ot = ot_desc_shelf.get((fibra_str, odf_calle_str, port_str))
        
        # Procesar PATCHEO
        if pd.isna(patcheo) or str(patcheo).strip() == '':
            sincroniza_base = "F (sin PATCHEO)"
            nuevas_filas.append({
                IS_COL_FIBRA: fibra,
                IS_COL_EQUIPO: equipo,
                IS_COL_INTERFAZ: interfaz,
                IS_COL_OP: None,
                IS_COL_CASSETERA: None,
                IS_COL_PATCHEO: patcheo,
                IS_COL_BANDEJA: None,
                IS_COL_ODF_CALLE: odf_calle,
                IS_COL_HILO: hilo,
                COL_DESC_PORT_OT: desc_port_ot,
                COL_DESC_SHELF_OT: desc_shelf_ot,
                IS_COL_ID_CABECERA: None,
                IS_COL_ID_PORT_PACHEO: None,
                COL_ID_PORT_ODF_TRONCAL: id_odf_troncal,
                'SINCRONIZA': sincroniza_base if id_odf_troncal is None or id_odf_troncal == '' else "F (sin PATCHEO)",
                IS_COL_ANILLO_ORIGEN: anillo_origen,
                'ANILLO': anillo,
                'CAMBIO': cambio
            })
            continue
        
        patcheo_str = str(patcheo).strip()
        if '-' not in patcheo_str:
            op = patcheo_str
            sincroniza_base = "F (formato PATCHEO sin guion)" if id_odf_troncal is None or id_odf_troncal == '' else "F"
            nuevas_filas.append({
                IS_COL_FIBRA: fibra,
                IS_COL_EQUIPO: equipo,
                IS_COL_INTERFAZ: interfaz,
                IS_COL_OP: op,
                IS_COL_CASSETERA: None,
                IS_COL_PATCHEO: patcheo,
                IS_COL_BANDEJA: None,
                IS_COL_ODF_CALLE: odf_calle,
                IS_COL_HILO: hilo,
                COL_DESC_PORT_OT: desc_port_ot,
                COL_DESC_SHELF_OT: desc_shelf_ot,
                IS_COL_ID_CABECERA: None,
                IS_COL_ID_PORT_PACHEO: None,
                COL_ID_PORT_ODF_TRONCAL: id_odf_troncal,
                'SINCRONIZA': sincroniza_base,
                IS_COL_ANILLO_ORIGEN: anillo_origen,
                'ANILLO': anillo,
                'CAMBIO': cambio
            })
            continue
        
        partes = patcheo_str.split('-', 1)
        op = partes[0].strip()
        resto = partes[1].strip() if len(partes) > 1 else ''
        if resto == '':
            sincroniza_base = "F (PATCHEO sin bandejas)" if id_odf_troncal is None or id_odf_troncal == '' else "F"
            nuevas_filas.append({
                IS_COL_FIBRA: fibra,
                IS_COL_EQUIPO: equipo,
                IS_COL_INTERFAZ: interfaz,
                IS_COL_OP: op,
                IS_COL_CASSETERA: None,
                IS_COL_PATCHEO: patcheo,
                IS_COL_BANDEJA: None,
                IS_COL_ODF_CALLE: odf_calle,
                IS_COL_HILO: hilo,
                COL_DESC_PORT_OT: desc_port_ot,
                COL_DESC_SHELF_OT: desc_shelf_ot,
                IS_COL_ID_CABECERA: None,
                IS_COL_ID_PORT_PACHEO: None,
                COL_ID_PORT_ODF_TRONCAL: id_odf_troncal,
                'SINCRONIZA': sincroniza_base,
                IS_COL_ANILLO_ORIGEN: anillo_origen,
                'ANILLO': anillo,
                'CAMBIO': cambio
            })
            continue
        
        bandejas_raw = [b.strip() for b in resto.split('/') if b.strip() != '']
        if not bandejas_raw:
            sincroniza_base = "F (PATCHEO sin bandejas)" if id_odf_troncal is None or id_odf_troncal == '' else "F"
            nuevas_filas.append({
                IS_COL_FIBRA: fibra,
                IS_COL_EQUIPO: equipo,
                IS_COL_INTERFAZ: interfaz,
                IS_COL_OP: op,
                IS_COL_CASSETERA: None,
                IS_COL_PATCHEO: patcheo,
                IS_COL_BANDEJA: None,
                IS_COL_ODF_CALLE: odf_calle,
                IS_COL_HILO: hilo,
                COL_DESC_PORT_OT: desc_port_ot,
                COL_DESC_SHELF_OT: desc_shelf_ot,
                IS_COL_ID_CABECERA: None,
                IS_COL_ID_PORT_PACHEO: None,
                COL_ID_PORT_ODF_TRONCAL: id_odf_troncal,
                'SINCRONIZA': sincroniza_base,
                IS_COL_ANILLO_ORIGEN: anillo_origen,
                'ANILLO': anillo,
                'CAMBIO': cambio
            })
            continue
        
        letra_comun = ''
        for b in bandejas_raw:
            match_letra = re.search(r'([A-Za-z])', b)
            if match_letra:
                letra_comun = match_letra.group(1).upper()
                break
        
        op_key = f"OP{equipo}_{op}" if equipo and op else None
        for bandeja_raw in bandejas_raw:
            match_num = re.search(r'(\d+)', bandeja_raw)
            num_bandeja = match_num.group(1) if match_num else ''
            
            id_port_pacheo = None
            if letra_comun and op_key and num_bandeja:
                clave_busqueda = (letra_comun, op_key, num_bandeja)
                id_port_pacheo = pacheo_dict.get(clave_busqueda)
            
            id_cabecera = None
            if id_backbone and id_port_pacheo:
                id_bay_backbone = id_a_id_bay.get(str(id_backbone))
                id_bay_pacheo = id_a_id_bay.get(str(id_port_pacheo))
                if id_bay_backbone is not None and id_bay_pacheo is not None and str(id_bay_backbone) == str(id_bay_pacheo):
                    id_cabecera = id_backbone
            
            sincroniza = "F"
            if id_odf_troncal is None or id_odf_troncal == '':
                sincroniza = "F (sin ID_PORT_ODF-TRONCAL)"
            else:
                sincroniza = "V"
                if equipo and interfaz:
                    clave = (equipo, interfaz)
                    if clave not in op_existe:
                        sincroniza = "F (sin registro en OP)"
                    elif not op_ids[clave]:
                        sincroniza = "F (ID vacío en OP)"
                    elif id_port_pacheo is None or str(id_port_pacheo) not in op_ids[clave]:
                        sincroniza = "F (ID no coincide[93XX-C12])"
                else:
                    sincroniza = "F (equipo o interfaz vacío)"
            
            nuevas_filas.append({
                IS_COL_FIBRA: fibra,
                IS_COL_EQUIPO: equipo,
                IS_COL_INTERFAZ: interfaz,
                IS_COL_OP: op,
                IS_COL_CASSETERA: letra_comun if letra_comun else None,
                IS_COL_PATCHEO: patcheo,
                IS_COL_BANDEJA: num_bandeja,
                IS_COL_ODF_CALLE: odf_calle,
                IS_COL_HILO: hilo,
                COL_DESC_PORT_OT: desc_port_ot,
                COL_DESC_SHELF_OT: desc_shelf_ot,
                IS_COL_ID_CABECERA: id_cabecera,
                IS_COL_ID_PORT_PACHEO: id_port_pacheo,
                COL_ID_PORT_ODF_TRONCAL: id_odf_troncal,
                'SINCRONIZA': sincroniza,
                IS_COL_ANILLO_ORIGEN: anillo_origen,
                'ANILLO': anillo,
                'CAMBIO': cambio
            })
    
    df_procesado = pd.DataFrame(nuevas_filas)
    df_procesado.columns = [c.upper() for c in df_procesado.columns]
    df_procesado = validar_anillos_por_fibra_y_hilo(df_procesado)
    
    column_order = [
        IS_COL_FIBRA,
        IS_COL_EQUIPO,
        IS_COL_INTERFAZ,
        IS_COL_OP,
        IS_COL_CASSETERA,
        IS_COL_PATCHEO,
        IS_COL_BANDEJA,
        IS_COL_ODF_CALLE,
        IS_COL_HILO,
        COL_DESC_PORT_OT,
        COL_DESC_SHELF_OT,
        IS_COL_ID_CABECERA,
        IS_COL_ID_PORT_PACHEO,
        COL_ID_PORT_ODF_TRONCAL,
        'SINCRONIZA',
        IS_COL_ANILLO_ORIGEN,
        'ANILLO',
        'CAMBIO'
    ]
    for col in column_order:
        if col not in df_procesado.columns:
            df_procesado[col] = None
    df_procesado = df_procesado[column_order]
    return df_procesado

def main():
    try:
        cables_omitidos = cargar_cables_omitidos(RUTA_CENTRALES)  # Solo para OM
        df_ec_nuevo = generar_hoja_ec(RUTA_BASE_EC)
        if RUTA_CENTRALES.exists():
            todas_hojas = pd.read_excel(RUTA_CENTRALES, sheet_name=None, dtype=str)
            for sheet_name in todas_hojas:
                if isinstance(todas_hojas[sheet_name], pd.DataFrame):
                    todas_hojas[sheet_name].columns = [c.upper() for c in todas_hojas[sheet_name].columns]
        else:
            todas_hojas = {}
        if not df_ec_nuevo.empty:
            todas_hojas[HOJA_ENTRADA_EC] = df_ec_nuevo
        elif HOJA_ENTRADA_EC not in todas_hojas:
            print(f"Error: no hay datos para la hoja {HOJA_ENTRADA_EC}", file=sys.stderr)
            return
        df_cabecera = todas_hojas[HOJA_ENTRADA_EC].copy()
        columnas_necesarias = [EC_COL_ARCHIVO, EC_COL_EQUIPO, EC_COL_PUERTO, EC_COL_ODF, EC_COL_CASSETERA, EC_COL_POS, EC_COL_POS2]
        for col in columnas_necesarias:
            if col not in df_cabecera.columns:
                print(f"Error: columna '{col}' no encontrada en la hoja {HOJA_ENTRADA_EC}", file=sys.stderr)
                return
        df_cabecera, df_op, df_ot = generar_hojas_op_y_ot(df_cabecera)
        df_is = generar_hoja_is(RUTA_BASE_IS, df_op, df_ot)
        df_is, _ = actualizar_ids_desde_ec(df_is, df_cabecera)
        columnas_ec_final = [EC_COL_ARCHIVO, EC_COL_EQUIPO, EC_COL_PUERTO, EC_COL_ODF, EC_COL_CASSETERA, EC_COL_POS, '_ODF_ORIGINAL', EC_COL_POS2, EC_COL_ID_PORT]
        columnas_ec_final = [c for c in columnas_ec_final if c in df_cabecera.columns]
        df_cabecera = df_cabecera[columnas_ec_final]
        columnas_op_final = [OP_COL_ARCHIVO, OP_COL_EQUIPO, OP_COL_PUERTO, OP_COL_ODF, OP_COL_CASSETERA, OP_COL_POS_ORIG, OP_COL_POS_TOKEN, OP_COL_ID_CABECERA, OP_COL_ID_PORT_PACHEO]
        columnas_op_final = [c for c in columnas_op_final if c in df_op.columns]
        df_op = df_op[columnas_op_final]
        todas_hojas[HOJA_ENTRADA_EC] = df_cabecera
        todas_hojas[HOJA_SALIDA_OP] = df_op
        if not df_ot.empty:
            todas_hojas[HOJA_SALIDA_OT] = df_ot
        if not df_is.empty:
            todas_hojas[HOJA_SALIDA_IS] = df_is
        if HOJA_OM not in todas_hojas:
            if cables_omitidos:
                df_om = pd.DataFrame(list(cables_omitidos), columns=['FIBRA'])
                todas_hojas[HOJA_OM] = df_om
        if "cabecera" in todas_hojas:
            del todas_hojas["cabecera"]
        with pd.ExcelWriter(RUTA_CENTRALES, engine='openpyxl') as writer:
            for sheet_name, df in todas_hojas.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    except Exception as e:
        print(f"Error en main: {e}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        raise

if __name__ == "__main__":
    main()